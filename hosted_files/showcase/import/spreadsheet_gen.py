#!/usr/bin/env python3
# Aria Corona February 25th, 2025

from pathlib import Path
import sys

ROOT_DIR = Path(__file__).resolve().parent.parent.parent
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.append(str(SRC_DIR))

from NotionApiHelper import NotionApiHelper
from AutomatedEmails import AutomatedEmails
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from loggers.logger import Logger
import json, openpyxl, os, logging

_notion = NotionApiHelper()

if __name__ == "__main__":
    # Logger Configuration
    _logger = Logger(
        log_name="SpreadsheetGen",
        info_file="SpreadsheetGen.log",
        error_file="SpreadsheetGen.log",
    )
    _logger = _logger.get_logger()
else:
    _logger = logging.getLogger(__name__)
    if not _logger.hasHandlers():  # Avoid adding duplicate handlers
        _logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
        )
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(formatter)
        _logger.addHandler(ch)

# Automated Email Configuration
automated_emails = AutomatedEmails()

# Alignment styles
_C_ALIGN = Alignment(horizontal="center", vertical="center")
_L_ALIGN = Alignment(horizontal="left", vertical="center")
_R_ALIGN = Alignment(horizontal="right", vertical="center")

# Font styles
_BOLD_FONT = Font(bold=True)

# Fill styles
_LIGHT_GRAY_FILL = PatternFill(
    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
)

# Border styles
_LTB_BORDER = Border(
    left=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)
_RTB_BORDER = Border(
    right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)
_RT_BORDER = Border(right=Side(style="thin"), top=Side(style="thin"))
_LT_BORDER = Border(left=Side(style="thin"), top=Side(style="thin"))
_TB_BORDER = Border(top=Side(style="thin"), bottom=Side(style="thin"))
_L_BORDER = Border(left=Side(style="thin"))
_R_BORDER = Border(right=Side(style="thin"))
_B_BORDER = Border(bottom=Side(style="thin"))
_LB_BORDER = Border(left=Side(style="thin"), bottom=Side(style="thin"))
_RB_BORDER = Border(right=Side(style="thin"), bottom=Side(style="thin"))


class ExportTable:
    """
    ExportTable is a class designed to query a database, parse the retrieved data, and export it to an Excel spreadsheet.
    It also provides functionality to email the generated report.
    Attributes:
        db_id (str): The ID of the database to query.
        filter (dict): The filter to apply when querying the database.
        property_names (list): The list of property names to extract from the database.
        output_file (str): The path to the output Excel file.
        email_conf_path (str): The path to the email configuration file.
        subject (str): The subject of the email.
        body (str): The body of the email.
        worksheet_title (str): The title of the worksheet in the Excel file.
        default_col_width (int): The default width of the columns in the Excel file.
        logger (Logger): The logger object for logging information.
    Methods:
        query_db(db_id, filter):
            Queries the database using the provided database ID and filter.
        parse_data(data):
        write_table(ws, data, headers, row_num):
        set_column_width(ws, data):
            Sets the width of the columns in the Excel worksheet based on the data.
        build_xlsx(parsed_data):
        run():
            Executes the process of exporting data to an Excel file.
        email_report():
            Sends an email with the generated Excel report as an attachment.
    """

    def __init__(
        self,
        db_id,
        query_filter,
        property_names,
        output_file,
        email_conf_path=None,
        subject=None,
        body=None,
        worksheet_title="untitled",
        default_col_width=15,
        logger=_logger,
    ):
        self.db_id = db_id
        self.filter = query_filter
        self.property_names = property_names
        self.output_file = output_file
        self.email_conf_path = email_conf_path
        self.subject = subject
        self.body = body
        self.worksheet_title = worksheet_title
        self.default_col_width = default_col_width
        self.logger = logger

        # Collects the largest width of each column based off of table header char count.
        self.max_width = {}

        if not os.path.exists(os.path.dirname(self.output_file)):
            os.makedirs(os.path.dirname(self.output_file))

        pass

    def query_db(self, db_id, filter):
        self.logger.info(f"Querying database {db_id}.")

        response = _notion.query(db_id, content_filter=filter)

        if not response:
            self.logger.error(f"Query failed for database {db_id}.")
            return []

        return response

    def parse_data(self, data):
        """
        Parses the given data and extracts specified properties.
        Args:
            data (list): A list of pages, where each page is a dictionary containing 'properties' and 'id'.
        Returns:
            list: A list of dictionaries, where each dictionary represents a parsed page with the specified properties.
            The keys are the property names and the values are the property values.
        """
        parsed_data = []

        for page in data:
            props = page["properties"]
            pid = page["id"]

            parsed_page = {}
            parsed_page["id"] = pid

            for prop_name in self.property_names:
                parsed_page[prop_name] = _notion.return_property_value(
                    props[prop_name], pid
                )
                if parsed_page[prop_name] is None:
                    parsed_page[prop_name] = ""

            parsed_data.append(parsed_page)

        return parsed_data

    def write_table(self, ws, data, headers, row_num):
        """
        Writes a table to an Excel worksheet, starting at the specified row number.
        Alternates the background color of rows for better readability, and adjusts column widths based on the data.
        Set up to handle any number of columns and rows greater than 2.
        Iterates through columns by converting a character to ASCII value and adding an offset.
        Args:
            ws (Worksheet): The worksheet object where the table will be written.
            data (dict): A dictionary containing the data to be written. The keys are customer names and the values are dictionaries of column data.
            total (int or float): The total value to be written at the end of the table.
            headers (list): A list of column headers.
            row_num (int): The starting row number for writing the table.
        Returns:
            tuple: A tuple containing the worksheet object and the next row number after the table.
        """

        self.logger.info("write_table() called.")

        # Convert char to ASCII value. Allows for iteration through columns.
        col_value = ord("A")

        # Get the length of the lists
        header_length = len(headers)
        data_length = len(data)

        # No data to write
        if not data:
            return ws, row_num

        # Write headers, degine header_row for alternating background color
        header_row = row_num
        for i, header in enumerate(headers):

            col = chr(col_value + i)
            ws[f"{col}{row_num}"] = header
            ws[f"{col}{row_num}"].font = _BOLD_FONT
            ws[f"{col}{row_num}"].alignment = _C_ALIGN
            ws[f"{col}{row_num}"].fill = _LIGHT_GRAY_FILL

            if i == 0:
                ws[f"{col}{row_num}"].border = _LTB_BORDER
            elif i == header_length - 1:
                ws[f"{col}{row_num}"].border = _RTB_BORDER
            else:
                ws[f"{col}{row_num}"].border = _TB_BORDER

        # Write data
        background_offset = header_row % 2
        for row_index, page in enumerate(data):
            row_num += 1

            product = page[headers[0]]

            col = chr(col_value)

            # start column A
            ws[f"{col}{row_num}"] = product
            ws[f"{col}{row_num}"].alignment = _C_ALIGN
            if row_index == data_length - 1:
                ws[f"{col}{row_num}"].border = _LB_BORDER
            else:
                ws[f"{col}{row_num}"].border = _L_BORDER

            # Alternates background color
            if row_num % 2 == background_offset:
                ws[f"{col}{row_num}"].fill = _LIGHT_GRAY_FILL

            # start column B-*
            for i, header in enumerate(headers):
                if i == 0:
                    continue

                col = chr(col_value + i)

                value = page[header]

                self.logger.debug(f"Row: {row_num}, Col: {col}, Value: {value}")
                ws[f"{col}{row_num}"] = value
                ws[f"{col}{row_num}"].alignment = _C_ALIGN

                # Alternates background color
                if row_num % 2 == background_offset:
                    ws[f"{col}{row_num}"].fill = _LIGHT_GRAY_FILL

                # Sets border for last column
                if i == header_length - 1 and row_index == data_length - 1:
                    ws[f"{col}{row_num}"].border = _RB_BORDER
                elif i == header_length - 1:
                    ws[f"{col}{row_num}"].border = _R_BORDER
                elif row_index == data_length - 1:
                    ws[f"{col}{row_num}"].border = _B_BORDER

        return ws, row_num

    def set_column_width(self, ws, data):
        col_value = ord("A")

        for page in data:
            for i, (name, value) in enumerate(page.items()):
                if i == 0:
                    continue

                col = chr(col_value + i - 1)

                if col not in self.max_width:
                    self.max_width[col] = self.default_col_width
                if len(str(value)) > self.max_width[col]:
                    self.logger.info(
                        f"Updating max width for column {col} to {len(str(value))}. Value: {value}"
                    )
                    self.max_width[col] = len(str(value))

        self.logger.debug(json.dumps(self.max_width, indent=4))

        for col, width in self.max_width.items():
            self.logger.debug(f"Setting column {col} width to {width + 5}.")
            ws.column_dimensions[col].width = width + 5

        return ws

    def build_xlsx(self, parsed_data):
        """
        Builds an Excel (.xlsx) file from the provided parsed data.
        Args:
            parsed_data (list of dict): The data to be written into the Excel file. Each dictionary in the list represents a row of data.
        Returns:
            None
        Side Effects:
            - Creates and saves an Excel file at the location specified by self.output_file.
            - Logs the process of building and saving the Excel file.
        Raises:
            Any exceptions raised by openpyxl during the creation or saving of the workbook.
        """
        self.logger.info("Building Excel file.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.worksheet_title

        row_number = 1
        ws, row_number = self.write_table(
            ws, parsed_data, self.property_names, row_number
        )
        ws = self.set_column_width(ws, parsed_data)

        self.logger.info("Saving Excel file.")
        wb.save(self.output_file)
        pass

    def run(self):
        """
        Executes the process of exporting data.
        This method performs the following steps:
        1. Queries the Notion database using the provided database ID and filter.
        2. Parses the retrieved Notion data.
        3. Builds an Excel spreadsheet from the parsed data.
        Returns:
            None
        """
        self.logger.info("Data Export started.")

        notion_data = self.query_db(self.db_id, self.filter)

        parsed_data = self.parse_data(notion_data)

        self.build_xlsx(parsed_data)

    def email_report(self):
        self.logger.info("Emailing report.")
        automated_emails.send_email(
            self.email_conf_path, self.subject, self.body, [self.output_file]
        )
        pass


class SpreadsheetGen:
    """
    A generic spreadsheet generator class for creating pre-formatted Excel files.

    Allows for flexible spreadsheet creation with custom headers, row-by-row data addition,
    and individual cell updates. The spreadsheet is maintained as a class attribute.

    Attributes:
        headers (list): List of column headers for the spreadsheet.
        worksheet_title (str): Title of the worksheet.
        logger (Logger): Logger object for logging information.
        workbook (Workbook): The openpyxl workbook object.
        worksheet (Worksheet): The active worksheet in the workbook.
        current_row (int): Tracks the current row number for adding data.

    Methods:
        add_row(data): Adds a complete row of data to the spreadsheet.
        set_cell(row, col, value): Sets a specific cell value.
        get_cell(row, col): Gets a specific cell value.
        apply_header_style(): Applies formatting to the header row.
        apply_alternating_rows(): Applies alternating background colors to data rows.
        auto_adjust_columns(): Automatically adjusts column widths based on content.
        save(file_path): Saves the spreadsheet to a file.
    """

    def __init__(
        self, headers, worksheet_title="Sheet1", auto_write_headers=True, logger=_logger
    ):
        """
        Initialize a new GenericSpreadsheet.

        Args:
            headers (list): List of column header names.
            worksheet_title (str): Title for the worksheet. Default is "Sheet1".
            auto_write_headers (bool): Whether to automatically write headers to row 1. Default is True.
            logger (Logger): Logger instance for logging. Default is module logger.
        """
        self.headers = headers
        self.worksheet_title = worksheet_title
        self.logger = logger

        # Create workbook and worksheet
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = worksheet_title

        # Track current row for sequential data addition
        self.current_row = 1

        # Write headers to first row (optional)
        if auto_write_headers:
            self._write_headers()

        self.logger.debug(
            f"GenericSpreadsheet initialized with {len(headers)} headers."
        )

    def _write_headers(self):
        """Write headers to the first row of the spreadsheet."""
        for col_idx, header in enumerate(self.headers, start=1):
            col_letter = get_column_letter(col_idx)
            cell = self.worksheet[f"{col_letter}{self.current_row}"]
            cell.value = header

        self.current_row += 1
        self.logger.debug(f"Headers written: {self.headers}")

    def add_row(self, data):
        """
        Add a complete row of data to the spreadsheet.

        Args:
            data (list or dict): Row data. If list, values are written in order.
                                If dict, keys should match header names.

        Returns:
            int: The row number where data was added.

        Raises:
            ValueError: If data doesn't match the number of headers or contains invalid keys.
        """
        if isinstance(data, dict):
            # Convert dict to list based on header order
            row_data = []
            for header in self.headers:
                if header not in data:
                    self.logger.warning(
                        f"Header '{header}' not found in data dict. Using empty string."
                    )
                    row_data.append("")
                else:
                    row_data.append(data[header])
        elif isinstance(data, (list, tuple)):
            row_data = list(data)
            if len(row_data) != len(self.headers):
                raise ValueError(
                    f"Data length ({len(row_data)}) doesn't match headers length ({len(self.headers)})"
                )
        else:
            raise ValueError(f"Data must be a list, tuple, or dict. Got {type(data)}")

        # Write data to current row
        for col_idx, value in enumerate(row_data, start=1):
            col_letter = get_column_letter(col_idx)
            self.worksheet[f"{col_letter}{self.current_row}"] = value

        added_row = self.current_row
        self.current_row += 1

        self.logger.debug(f"Row {added_row} added with {len(row_data)} values.")
        return added_row

    def set_cell(self, row, col, value):
        """
        Set a specific cell value in the spreadsheet.

        Args:
            row (int): Row number (1-indexed).
            col (int or str): Column number (1-indexed) or column letter (e.g., 'A', 'B').
            value: Value to set in the cell.

        Returns:
            str: The cell reference (e.g., 'A1').
        """
        if isinstance(col, int):
            col_letter = get_column_letter(col)
        else:
            col_letter = col.upper()

        cell_ref = f"{col_letter}{row}"
        self.worksheet[cell_ref] = value

        self.logger.debug(f"Cell {cell_ref} set to: {value}")
        return cell_ref

    def get_cell(self, row, col):
        """
        Get a specific cell value from the spreadsheet.

        Args:
            row (int): Row number (1-indexed).
            col (int or str): Column number (1-indexed) or column letter.

        Returns:
            The value of the cell.
        """
        if isinstance(col, int):
            col_letter = get_column_letter(col)
        else:
            col_letter = col.upper()

        cell_ref = f"{col_letter}{row}"
        value = self.worksheet[cell_ref].value

        return value

    def apply_header_style(self, bold=True, bg_color="D3D3D3", alignment="center"):
        """
        Apply formatting to the header row.

        Args:
            bold (bool): Whether to make headers bold. Default is True.
            bg_color (str): Hex color code for background. Default is light gray.
            alignment (str): Text alignment ('center', 'left', 'right'). Default is 'center'.
        """
        align_map = {"center": _C_ALIGN, "left": _L_ALIGN, "right": _R_ALIGN}

        alignment_style = align_map.get(alignment.lower(), _C_ALIGN)
        font_style = _BOLD_FONT if bold else Font()
        fill_style = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )

        for col_idx in range(1, len(self.headers) + 1):
            col_letter = get_column_letter(col_idx)
            cell = self.worksheet[f"{col_letter}1"]
            cell.font = font_style
            cell.alignment = alignment_style
            cell.fill = fill_style

            # Add borders
            if col_idx == 1:
                cell.border = _LTB_BORDER
            elif col_idx == len(self.headers):
                cell.border = _RTB_BORDER
            else:
                cell.border = _TB_BORDER

        self.logger.debug("Header styling applied.")

    def apply_alternating_rows(self, start_row=2, bg_color="D3D3D3"):
        """
        Apply alternating background colors to data rows.

        Args:
            start_row (int): Row number to start alternating colors. Default is 2 (after headers).
            bg_color (str): Hex color code for alternating rows. Default is light gray.
        """
        fill_style = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )

        for row_idx in range(start_row, self.current_row):
            if row_idx % 2 == 0:  # Even rows get the background color
                for col_idx in range(1, len(self.headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    cell = self.worksheet[f"{col_letter}{row_idx}"]
                    cell.fill = fill_style

        self.logger.debug(f"Alternating row colors applied from row {start_row}.")

    def auto_adjust_columns(self, min_width=10, max_width=50, padding=2):
        """
        Automatically adjust column widths based on content.

        Args:
            min_width (int): Minimum column width. Default is 10.
            max_width (int): Maximum column width. Default is 50.
            padding (int): Extra padding to add to calculated width. Default is 2.
        """
        for col_idx in range(1, len(self.headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            # Check all cells in the column
            for row_idx in range(1, self.current_row):
                cell_value = self.worksheet[f"{col_letter}{row_idx}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set adjusted width
            adjusted_width = min(max(max_length + padding, min_width), max_width)
            self.worksheet.column_dimensions[col_letter].width = adjusted_width

        self.logger.debug("Column widths auto-adjusted.")

    def apply_borders(self, start_row=1, end_row=None):
        """
        Apply borders to a range of cells.

        Args:
            start_row (int): Starting row for borders. Default is 1.
            end_row (int): Ending row for borders. If None, uses current_row - 1.
        """
        if end_row is None:
            end_row = self.current_row - 1

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, len(self.headers) + 1):
                col_letter = get_column_letter(col_idx)
                cell = self.worksheet[f"{col_letter}{row_idx}"]
                cell.border = thin_border

        self.logger.debug(f"Borders applied from row {start_row} to {end_row}.")

    def save(self, file_path):
        """
        Save the spreadsheet to a file.

        Args:
            file_path (str): Full path where the file should be saved.

        Raises:
            OSError: If the directory doesn't exist or file can't be written.
        """
        # Ensure directory exists
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
            self.logger.info(f"Created directory: {directory}")

        # Save workbook
        self.workbook.save(file_path)
        self.logger.info(f"Spreadsheet saved to: {file_path}")

    def get_row_count(self):
        """
        Get the current number of rows (including headers).

        Returns:
            int: Number of rows with data.
        """
        return self.current_row - 1

    def get_column_count(self):
        """
        Get the number of columns.

        Returns:
            int: Number of columns.
        """
        return len(self.headers)

    def add_worksheet(self, title, headers=None, auto_write_headers=True):
        """
        Add a new worksheet to the workbook and switch to it.

        Args:
            title (str): Title for the new worksheet.
            headers (list): Optional list of headers for the new worksheet.
                           If None, uses the same headers as the original sheet.
            auto_write_headers (bool): Whether to automatically write headers to row 1. Default is True.

        Returns:
            Worksheet: The newly created worksheet.
        """
        # Create new worksheet
        new_worksheet = self.workbook.create_sheet(title=title)

        # Switch to new worksheet
        self.worksheet = new_worksheet

        # Update headers if provided, otherwise use existing
        if headers is not None:
            self.headers = headers

        # Reset row counter
        self.current_row = 1

        # Write headers (optional)
        if auto_write_headers:
            self._write_headers()

        self.logger.debug(f"New worksheet '{title}' created and activated.")
        return new_worksheet

    def switch_worksheet(self, title):
        """
        Switch to an existing worksheet by title.

        Args:
            title (str): Title of the worksheet to switch to.

        Raises:
            ValueError: If worksheet with the given title doesn't exist.
        """
        for sheet in self.workbook.worksheets:
            if sheet.title == title:
                self.worksheet = sheet
                # Update current_row to the next available row
                self.current_row = sheet.max_row + 1
                self.logger.debug(f"Switched to worksheet '{title}'.")
                return

        raise ValueError(f"Worksheet '{title}' not found in workbook.")

    def get_worksheet_names(self):
        """
        Get a list of all worksheet names in the workbook.

        Returns:
            list: List of worksheet title strings.
        """
        return [sheet.title for sheet in self.workbook.worksheets]
